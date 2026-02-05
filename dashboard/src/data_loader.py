import pandas as pd
import os
from openpyxl import load_workbook

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)

KSE_FILE = os.path.join(DATA_DIR, "KSE100_Monthly_Investment.xlsx")
KMI_FILE = os.path.join(DATA_DIR, "KMI30_Monthly_Investment.xlsx")
CACHE_FILE = os.path.join(DATA_DIR, "Common_Stocks_Cache.xlsx")

def save_scraped_data(kse_df, kmi_df):
    """
    Saves scraped data to Excel files.
    """
    if kse_df is not None:
        kse_df.to_excel(KSE_FILE, index=False)
    
    if kmi_df is not None:
        kmi_df.to_excel(KMI_FILE, index=False)

def load_data_from_file():
    """
    Loads data from existing Excel files.
    """
    kse_df = None
    kmi_df = None
    
    if os.path.exists(KSE_FILE):
        try:
            kse_df = pd.read_excel(KSE_FILE)
        except Exception as e:
            print(f"Error loading KSE file: {e}")
            
    if os.path.exists(KMI_FILE):
        try:
            kmi_df = pd.read_excel(KMI_FILE)
        except Exception as e:
            print(f"Error loading KMI file: {e}")
            
    return kse_df, kmi_df

def save_common_stocks(df, total_investment):
    """
    Saves the final calculation to Excel.
    """
    output_path = os.path.join(DATA_DIR, "Final_Allocation.xlsx")
    
    # Save simple dataframe first
    df.to_excel(output_path, index=False, sheet_name="Allocation")
    
    # Add metadata
    # We could use openpyxl to add the 'Total Investment' header as requested in the prompt
    # "Export final allocation to Excel"
    try:
        wb = load_workbook(output_path)
        ws = wb["Allocation"]
        ws.insert_rows(1, amount=3)
        ws["A1"] = "Total Monthly Investment"
        ws["B1"] = total_investment
        wb.save(output_path)
        return output_path
    except Exception as e:
        print(f"Error formatting excel: {e}")
        return output_path
